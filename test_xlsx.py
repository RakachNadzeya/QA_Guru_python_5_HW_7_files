import os
import xlrd as xlrd
from openpyxl import load_workbook
from conftest import PROJECT_ROOT_PATH


def test_read_xlsx_file():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    xlsx_file_path = os.path.join(PROJECT_ROOT_PATH, 'resources', 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    assert sheet.cell(row=3, column=2).value == 'Mara'
